from http.server import BaseHTTPRequestHandler
import json
import io

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint


# ── Palette ──────────────────────────────────────────────────────────────────
TEAL       = "0D9488"
TEAL_SOFT  = "F0FDFA"
ORANGE     = "F97316"
ORANGE_SOFT= "FFF7ED"
PURPLE     = "7C3AED"
PURPLE_SOFT= "F5F3FF"
BLUE       = "3B82F6"
BLUE_SOFT  = "EFF6FF"
DARK       = "1A1A2E"
GRAY       = "6B7280"
LGRAY      = "F1F5F9"
WHITE      = "FFFFFF"
GREEN      = "22C55E"
GREEN_SOFT = "F0FDF4"

PIE_COLORS = [TEAL, ORANGE, PURPLE, BLUE, GREEN,
              "F43F5E", "EAB308", "06B6D4", "8B5CF6", "EC4899",
              "14B8A6", "F59E0B", "6366F1", "10B981", "EF4444"]

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=DARK, size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def left():
    return Alignment(horizontal="left", vertical="center")

def right_align():
    return Alignment(horizontal="right", vertical="center")

def thin_border():
    s = Side(style="thin", color="E5E7EB")
    return Border(left=s, right=s, top=s, bottom=s)

def fmt_eur(ws, cell_ref):
    ws[cell_ref].number_format = '#,##0 "€"'

def fmt_pct(ws, cell_ref):
    ws[cell_ref].number_format = '0.0%'

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def row_height(ws, row, h):
    ws.row_dimensions[row].height = h


# ── Feuille récapitulatif ─────────────────────────────────────────────────────
def build_recap(wb, results):
    ws = wb.create_sheet("Récapitulatif", 0)
    ws.sheet_view.showGridLines = False

    # Titre
    ws.merge_cells("A1:F1")
    ws["A1"] = "📊  Analyse des Offres"
    ws["A1"].font = Font(bold=True, size=16, color=WHITE, name="Calibri")
    ws["A1"].fill = fill(TEAL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    row_height(ws, 1, 36)

    # Sous-titre date
    from datetime import date
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Exporté le {date.today().strftime('%d/%m/%Y')} — CA Hors Taxes"
    ws["A2"].font = Font(italic=True, size=10, color=GRAY, name="Calibri")
    ws["A2"].fill = fill(TEAL_SOFT)
    ws["A2"].alignment = center()
    row_height(ws, 2, 20)

    ws.append([])  # ligne vide
    row_height(ws, 3, 6)

    # En-têtes
    headers = ["Code Offre", "Libellé", "CA HT Total", "Qté Vendue", "Commandes", "Délégués"]
    ws.append(headers)
    row_height(ws, 4, 24)
    for col, h in enumerate(headers, 1):
        c = ws.cell(4, col)
        c.font = font(bold=True, color=WHITE, size=10)
        c.fill = fill(DARK)
        c.alignment = center()
        c.border = thin_border()

    # Lignes données
    for i, r in enumerate(results):
        row = 5 + i
        values = [
            r.get("offre", {}).get("code", ""),
            r.get("offre", {}).get("label", "") or "",
            r.get("caTotal", 0),
            r.get("qtyTotal", 0),
            len(r.get("debugOrders", [])),
            len(r.get("delegues", [])),
        ]
        ws.append(values)
        row_height(ws, row, 22)
        bg = WHITE if i % 2 == 0 else LGRAY
        for col in range(1, 7):
            c = ws.cell(row, col)
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = center() if col > 2 else left()
            c.font = font(size=10)
        # CA en vert si > 0
        ca_cell = ws.cell(row, 3)
        ca_cell.number_format = '#,##0 "€"'
        ca_cell.font = font(bold=True, color=TEAL, size=11)

    # Total
    n = len(results)
    total_row = 5 + n
    ws.append(["TOTAL", "", f"=SUM(C5:C{total_row-1})", f"=SUM(D5:D{total_row-1})", "", ""])
    row_height(ws, total_row, 26)
    for col in range(1, 7):
        c = ws.cell(total_row, col)
        c.fill = fill(TEAL)
        c.font = font(bold=True, color=WHITE, size=11)
        c.alignment = center()
        c.border = thin_border()
    ws.cell(total_row, 3).number_format = '#,##0 "€"'

    # Largeurs
    for col, w in zip(range(1, 7), [14, 28, 16, 14, 12, 12]):
        set_col_width(ws, col, w)


# ── Feuille par offre ──────────────────────────────────────────────────────────
def build_offre_sheet(wb, r):
    code    = r.get("offre", {}).get("code", "Offre")
    label   = r.get("offre", {}).get("label", "")
    ca      = r.get("caTotal", 0)
    qty     = r.get("qtyTotal", 0)
    prods   = r.get("produits", [])
    delegs  = r.get("delegues", [])
    orders  = r.get("debugOrders", [])

    sheet_name = str(code)[:31]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # ── Bandeau titre ──
    ws.merge_cells("A1:G1")
    ws["A1"] = f"🏷  Offre  {code}" + (f"  —  {label}" if label else "")
    ws["A1"].font = Font(bold=True, size=15, color=WHITE, name="Calibri")
    ws["A1"].fill = fill(TEAL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    row_height(ws, 1, 38)

    # ── KPIs ──
    kpi_row = 2
    ws.merge_cells(f"A{kpi_row}:B{kpi_row}")
    ws[f"A{kpi_row}"] = "CA HT Total"
    ws[f"A{kpi_row}"].font = font(bold=True, color=TEAL, size=10)
    ws[f"A{kpi_row}"].fill = fill(TEAL_SOFT)
    ws[f"A{kpi_row}"].alignment = center()

    ws.merge_cells(f"C{kpi_row}:D{kpi_row}")
    ws[f"C{kpi_row}"] = ca
    ws[f"C{kpi_row}"].font = font(bold=True, color=TEAL, size=16)
    ws[f"C{kpi_row}"].fill = fill(TEAL_SOFT)
    ws[f"C{kpi_row}"].alignment = center()
    ws[f"C{kpi_row}"].number_format = '#,##0 "€"'

    ws[f"E{kpi_row}"] = "Qté vendue"
    ws[f"E{kpi_row}"].font = font(bold=True, color=ORANGE, size=10)
    ws[f"E{kpi_row}"].fill = fill(ORANGE_SOFT)
    ws[f"E{kpi_row}"].alignment = center()

    ws[f"F{kpi_row}"] = qty
    ws[f"F{kpi_row}"].font = font(bold=True, color=ORANGE, size=16)
    ws[f"F{kpi_row}"].fill = fill(ORANGE_SOFT)
    ws[f"F{kpi_row}"].alignment = center()

    ws[f"G{kpi_row}"] = f"{len(orders)} commandes"
    ws[f"G{kpi_row}"].font = font(bold=True, color=GRAY, size=10)
    ws[f"G{kpi_row}"].fill = fill(LGRAY)
    ws[f"G{kpi_row}"].alignment = center()

    row_height(ws, kpi_row, 36)

    # Ligne vide
    row_height(ws, 3, 8)

    # ── Tableau produits ──
    prod_start = 4
    ws.merge_cells(f"A{prod_start}:E{prod_start}")
    ws[f"A{prod_start}"] = "📦  Produits composants"
    ws[f"A{prod_start}"].font = font(bold=True, color=WHITE, size=11)
    ws[f"A{prod_start}"].fill = fill(BLUE)
    ws[f"A{prod_start}"].alignment = center()
    row_height(ws, prod_start, 24)

    hdr_row = prod_start + 1
    for col, h in enumerate(["Référence", "Nom produit", "Qté vendue", "CA HT (€)", "% CA"], 1):
        c = ws.cell(hdr_row, col)
        c.value = h
        c.font = font(bold=True, color=WHITE, size=10)
        c.fill = fill(BLUE_SOFT[:-2] + "BB")  # teinte légère
        c.fill = fill("BFDBFE")
        c.alignment = center()
        c.border = thin_border()
        c.font = font(bold=True, color="1D4ED8", size=10)
    row_height(ws, hdr_row, 20)

    data_start = hdr_row + 1
    for i, p in enumerate(prods):
        row = data_start + i
        pca = p.get("ca", 0)
        pct = pca / ca if ca > 0 else 0
        bg = WHITE if i % 2 == 0 else BLUE_SOFT
        vals = [p.get("ref", ""), p.get("name", ""), p.get("qtyVendue", 0), pca, pct]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row, col)
            c.value = val
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = left() if col <= 2 else center()
            c.font = font(size=10)
        ws.cell(row, 4).number_format = '#,##0 "€"'
        ws.cell(row, 5).number_format = '0.0%'
        row_height(ws, row, 20)

    # Ligne total produits
    tot_row = data_start + len(prods)
    for col in range(1, 6):
        c = ws.cell(tot_row, col)
        c.fill = fill(TEAL)
        c.font = font(bold=True, color=WHITE, size=10)
        c.alignment = center()
        c.border = thin_border()
    ws.cell(tot_row, 1).value = "TOTAL"
    ws.cell(tot_row, 3).value = f"=SUM(C{data_start}:C{tot_row-1})"
    ws.cell(tot_row, 4).value = f"=SUM(D{data_start}:D{tot_row-1})"
    ws.cell(tot_row, 4).number_format = '#,##0 "€"'
    row_height(ws, tot_row, 22)

    # ── Camembert produits — ancré à G4, taille fixe ──
    if len(prods) >= 2:
        pie_prod = PieChart()
        pie_prod.title = f"CA par produit — {code}"
        pie_prod.style = 10
        pie_prod.dataLabels = None

        data_ref = Reference(ws, min_col=4, min_row=data_start, max_row=data_start + len(prods) - 1)
        cats_ref = Reference(ws, min_col=1, min_row=data_start, max_row=data_start + len(prods) - 1)
        pie_prod.add_data(data_ref)
        pie_prod.set_categories(cats_ref)
        pie_prod.series[0].title = None

        for j in range(len(prods)):
            pt = DataPoint(idx=j)
            pt.graphicalProperties.solidFill = PIE_COLORS[j % len(PIE_COLORS)]
            pie_prod.series[0].dPt.append(pt)

        pie_prod.width  = 14
        pie_prod.height = 13
        ws.add_chart(pie_prod, "G4")   # ← position fixe

    # Ligne vide après produits
    after_prod = tot_row + 1
    row_height(ws, after_prod, 8)

    # ── Tableau délégués ──
    if delegs:
        del_start = after_prod + 1
        ws.merge_cells(f"A{del_start}:E{del_start}")
        ws[f"A{del_start}"] = "👤  Par délégué"
        ws[f"A{del_start}"].font = font(bold=True, color=WHITE, size=11)
        ws[f"A{del_start}"].fill = fill(PURPLE)
        ws[f"A{del_start}"].alignment = center()
        row_height(ws, del_start, 24)

        dhdr = del_start + 1
        for col, h in enumerate(["Délégué", "Qté vendue", "CA HT (€)", "% CA"], 1):
            c = ws.cell(dhdr, col)
            c.value = h
            c.font = font(bold=True, color=PURPLE, size=10)
            c.fill = fill("EDE9FE")
            c.alignment = center()
            c.border = thin_border()
        row_height(ws, dhdr, 20)

        dd_start = dhdr + 1
        for i, d in enumerate(delegs):
            row = dd_start + i
            dca = d.get("ca", 0)
            pct = dca / ca if ca > 0 else 0
            bg = WHITE if i % 2 == 0 else PURPLE_SOFT
            vals = [d.get("name", ""), d.get("qtyVendue", 0), dca, pct]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row, col)
                c.value = val
                c.fill = fill(bg)
                c.border = thin_border()
                c.alignment = left() if col == 1 else center()
                c.font = font(size=10)
            ws.cell(row, 3).number_format = '#,##0 "€"'
            ws.cell(row, 4).number_format = '0.0%'
            row_height(ws, row, 20)

        # Total délégués
        dtot = dd_start + len(delegs)
        for col in range(1, 5):
            c = ws.cell(dtot, col)
            c.fill = fill(PURPLE)
            c.font = font(bold=True, color=WHITE, size=10)
            c.alignment = center()
            c.border = thin_border()
        ws.cell(dtot, 1).value = "TOTAL"
        ws.cell(dtot, 3).value = f"=SUM(C{dd_start}:C{dtot-1})"
        ws.cell(dtot, 3).number_format = '#,##0 "€"'
        row_height(ws, dtot, 22)

        # Camembert délégués — ancré à N4 (à droite du camembert produits)
        if len(delegs) >= 2:
            pie_del = PieChart()
            pie_del.title = f"CA par délégué — {code}"
            pie_del.style = 10

            dr = Reference(ws, min_col=3, min_row=dd_start, max_row=dd_start + len(delegs) - 1)
            cr = Reference(ws, min_col=1, min_row=dd_start, max_row=dd_start + len(delegs) - 1)
            pie_del.add_data(dr)
            pie_del.set_categories(cr)
            pie_del.series[0].title = None

            for j in range(len(delegs)):
                pt = DataPoint(idx=j)
                pt.graphicalProperties.solidFill = PIE_COLORS[j % len(PIE_COLORS)]
                pie_del.series[0].dPt.append(pt)

            pie_del.width  = 14
            pie_del.height = 13
            ws.add_chart(pie_del, "N4")  # ← côte à côte avec le camembert produits

    # ── Largeurs colonnes ──
    for col, w in zip(range(1, 22), [14, 36, 14, 16, 10, 3, 2, 12, 12, 12, 12, 12, 3, 12, 12, 12, 12, 12, 12, 12, 12]):
        set_col_width(ws, col, w)


# ── Handler HTTP ──────────────────────────────────────────────────────────────
class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_POST(self):
        length  = int(self.headers.get("Content-Length", 0))
        payload = json.loads(self.rfile.read(length))
        results = payload.get("results", [])

        # Filtrer ceux qui ont des données
        valid = [r for r in results if not r.get("loading") and not r.get("error")]

        wb = Workbook()
        wb.remove(wb.active)  # supprimer la feuille vide

        build_recap(wb, valid)
        for r in valid:
            build_offre_sheet(wb, r)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        data = buf.read()

        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", "attachment; filename=analyse_offres.xlsx")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, *args):
        pass
